import importlib.util
import os
from pathlib import Path

import mne
import numpy as np
import pytest
from openpyxl import load_workbook

from src.hfo_app import HFO_App
from src.param.param_detector import ParamDetector, ParamHIL, ParamMNI, ParamSTE, ParamYASA
from src.param.param_filter import ParamFilter, ParamFilterSpindle
from src.spindle_app import SpindleApp


if os.environ.get("PYHFO_RUN_SMOKE") != "1":
    pytest.skip("backend smoke tests are opt-in; set PYHFO_RUN_SMOKE=1 to run them", allow_module_level=True)


ROOT = Path(__file__).resolve().parents[1]
SAMPLE_EDF = ROOT / "SM_B_ave.edf"
EXPECTED_EVENTS_XLSX = ROOT / "SM_B_ave.xlsx"


@pytest.fixture(scope="session")
def expected_hfo_events():
    workbook = load_workbook(EXPECTED_EVENTS_XLSX, read_only=True, data_only=True)
    sheet = workbook["Events"]
    return [
        (str(channel_name), int(start), int(end))
        for channel_name, start, end, *_rest in sheet.iter_rows(min_row=2, values_only=True)
        if channel_name is not None
    ]


@pytest.fixture(scope="session")
def hfo_cropped_raw():
    raw = mne.io.read_raw_edf(SAMPLE_EDF, preload=True, verbose=False)
    raw.crop(tmin=0, tmax=40)
    raw.pick(raw.ch_names[:8])
    return raw


def _read_sheet_rows(path, sheet_name):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    return list(sheet.iter_rows(values_only=True))


def _count_feature_images(folder):
    return len(list(Path(folder).glob("**/*.jpg")))


def _run_sample_ste_detection():
    app = HFO_App()
    app.set_n_jobs(max(1, min(4, os.cpu_count() or 1)))
    app.load_edf(str(SAMPLE_EDF))
    app.set_filter_parameter(ParamFilter())
    app.filter_eeg_data()
    app.set_detector(
        ParamDetector(
            ParamSTE(sample_freq=app.sample_freq, n_jobs=app.n_jobs),
            detector_type="STE",
        )
    )
    app.detect_biomarker()
    return app


def _run_hfo_detector(app, detector_type):
    detector_params = {
        "STE": ParamSTE,
        "MNI": ParamMNI,
        "HIL": ParamHIL,
    }[detector_type]
    app.set_detector(
        ParamDetector(
            detector_params(sample_freq=app.sample_freq, n_jobs=app.n_jobs),
            detector_type=detector_type,
        )
    )
    app.detect_biomarker()
    return app.analysis_session.active_run_id, app.event_features.get_num_biomarker()


def _build_line_noise_raw():
    sfreq = 200.0
    samples = int(sfreq * 2)
    times = np.arange(samples) / sfreq
    info = mne.create_info(["A1Ref", "A2Ref", "B10Ref"], sfreq, ch_types="eeg")
    line_noise = np.sin(2 * np.pi * 60 * times)
    data = np.vstack(
        [
            0.8 * np.sin(2 * np.pi * 8 * times) + 0.4 * line_noise,
            0.6 * np.cos(2 * np.pi * 12 * times) + 0.35 * line_noise,
            0.5 * np.sin(2 * np.pi * 16 * times) + 0.3 * line_noise,
        ]
    ) * 1e-6
    raw = mne.io.RawArray(data, info, verbose=False)
    raw.info["line_freq"] = 60
    return raw


def _build_spindle_recording(tmp_path):
    sfreq = 2000.0
    seconds = 30
    times = np.arange(int(sfreq * seconds)) / sfreq

    signal = np.random.default_rng(0).normal(scale=5e-7, size=times.shape)
    for start in [5, 12, 20]:
        mask = (times >= start) & (times < start + 1.0)
        envelope = np.hanning(mask.sum())
        signal[mask] += 2e-5 * np.sin(2 * np.pi * 13 * times[mask]) * envelope

    reference = np.random.default_rng(1).normal(scale=4e-7, size=times.shape)
    info = mne.create_info(["C3", "C4"], sfreq, ch_types="eeg")
    raw = mne.io.RawArray(np.vstack([signal, reference]), info, verbose=False)
    raw.info["line_freq"] = 60

    path = tmp_path / "synthetic_spindle_raw_eeg.fif"
    raw.save(path, overwrite=True, verbose=False)
    return path


def test_hfo_backend_signal_preparation_smoke():
    app = HFO_App()
    app.load_raw(_build_line_noise_raw(), file_path="<synthetic>")

    original_eeg = app.eeg_data.copy()
    filtered_60 = app.ensure_eeg_data_60()

    assert filtered_60.shape == original_eeg.shape
    assert not np.allclose(filtered_60, original_eeg)

    app.add_bipolar_channel("A1Ref", "A2Ref")
    assert app.channel_names[0] == "A1Ref#-#A2Ref"
    assert app.get_eeg_data_shape()[0] == 4

    app.set_filter_parameter(ParamFilter(fp=80, fs=90, sample_freq=app.sample_freq))
    app.filter_eeg_data()
    assert app.filtered is True
    assert app.has_filtered_data() is True

    app.set_filter_60()
    assert app.filter_data.shape == app.filter_data_60.shape
    app.set_unfiltered_60()
    np.testing.assert_allclose(app.eeg_data, app.eeg_data_un60)
    np.testing.assert_allclose(app.filter_data, app.filter_data_un60)


def test_hfo_backend_full_sample_ste_reference_smoke(tmp_path, expected_hfo_events):
    app = _run_sample_ste_detection()

    assert app.sample_freq == 2000
    assert app.eeg_data.shape == (40, 1_306_000)
    assert app.filtered is True
    assert app.detected is True
    assert app.event_features is not None

    detected_events = list(
        zip(
            app.event_features.channel_names.tolist(),
            app.event_features.starts.tolist(),
            app.event_features.ends.tolist(),
        )
    )
    assert detected_events == expected_hfo_events
    assert app.event_features.get_num_biomarker() == len(expected_hfo_events) == 583

    summaries = app.get_run_summaries()
    assert len(summaries) == 1
    assert summaries[0]["detector_name"] == "STE"
    assert summaries[0]["num_events"] == 583
    assert summaries[0]["num_channels"] == 40

    accepted = app.accept_active_run()
    assert accepted is not None
    decision = app.get_decision_summary()
    assert decision["num_runs"] == 1
    assert decision["active_detector"] == "STE"
    assert decision["accepted_detector"] == "STE"
    assert decision["top_channel"] is not None

    clinical_summary_path = tmp_path / "smoke_clinical_summary.xlsx"
    app.export_clinical_summary(clinical_summary_path)
    clinical_events = _read_sheet_rows(clinical_summary_path, "Active Run Events")
    assert len(clinical_events) == 584
    assert clinical_events[1][:3] == expected_hfo_events[0]

    workbook_export_path = tmp_path / "smoke_events.xlsx"
    app.export_excel(workbook_export_path)
    workbook_events = _read_sheet_rows(workbook_export_path, "Events")
    assert len(workbook_events) == 584
    assert workbook_events[1][:3] == expected_hfo_events[0]

    session_path = tmp_path / "smoke_session.pybrain"
    app.export_app(session_path)
    restored = HFO_App.import_app(session_path)
    restored_summaries = restored.get_run_summaries()
    assert len(restored_summaries) == 1
    assert restored_summaries[0]["detector_name"] == "STE"
    assert restored_summaries[0]["num_events"] == 583
    assert restored.get_decision_summary()["accepted_detector"] == "STE"


def test_hfo_backend_multi_detector_classification_and_export_smoke(hfo_cropped_raw, tmp_path):
    if importlib.util.find_spec("torch") is None:
        pytest.skip("torch is required for real artifact/spike/ehfo classification smoke")

    app = HFO_App()
    app.set_n_jobs(2)
    app.load_raw(hfo_cropped_raw.copy(), file_path="<cropped-smoke>")
    app.set_filter_parameter(ParamFilter())
    app.filter_eeg_data()

    ste_run_id, ste_events = _run_hfo_detector(app, "STE")
    mni_run_id, mni_events = _run_hfo_detector(app, "MNI")

    assert ste_events > 0
    assert mni_events > 0
    assert ste_run_id != mni_run_id
    assert len(app.get_run_summaries()) == 2

    app.set_default_cpu_classifier()
    app.classify_artifacts([0.0, 0.0], 0.5)
    app.classify_spikes()
    app.classify_ehfos()

    assert app.classified is True
    assert app.event_features.has_feature() is True
    assert app.event_features.artifact_predicted is True
    assert app.event_features.spike_predicted is True
    assert app.event_features.ehfo_predicted is True
    assert app.event_features.get_num_artifact() + app.event_features.get_num_real() == mni_events
    assert len(app.event_features.artifact_predictions) == mni_events
    assert len(app.event_features.spike_predictions) == mni_events
    assert len(app.event_features.ehfo_predictions) == mni_events

    accepted_run = app.accept_active_run()
    assert accepted_run is not None
    assert accepted_run.run_id == mni_run_id

    report_csv_path = tmp_path / "cropped_report.csv"
    app.export_report(report_csv_path)
    assert report_csv_path.exists()
    assert "ehfo" in report_csv_path.read_text(encoding="utf-8").splitlines()[0].lower()

    features_dir = tmp_path / "cropped_features"
    app.export_features(features_dir)
    assert (features_dir / "artifact").exists()
    assert (features_dir / "spike").exists()
    assert (features_dir / "non_spike").exists()
    assert _count_feature_images(features_dir) == mni_events

    workbook_export_path = tmp_path / "cropped_events.xlsx"
    app.export_excel(workbook_export_path)
    workbook_events = _read_sheet_rows(workbook_export_path, "Events")
    assert len(workbook_events) == mni_events + 1
    assert "eHFO" in workbook_events[0]

    clinical_summary_path = tmp_path / "cropped_clinical_summary.xlsx"
    app.export_clinical_summary(clinical_summary_path)
    decision_rows = _read_sheet_rows(clinical_summary_path, "Decision")
    assert decision_rows[1][2] == "MNI"

    hil_run_id, hil_events = _run_hfo_detector(app, "HIL")
    assert hil_events > 0
    assert hil_run_id not in {ste_run_id, mni_run_id}

    summaries = app.get_run_summaries()
    detectors = {summary["detector_name"] for summary in summaries}
    assert detectors == {"STE", "MNI", "HIL"}
    assert len(summaries) == 3

    app.set_run_visible(mni_run_id, False)
    assert all(run.run_id != mni_run_id for run in app.get_visible_runs())
    app.set_run_visible(mni_run_id, True)
    assert any(run.run_id == mni_run_id for run in app.get_visible_runs())

    comparison = app.compare_runs()
    assert len(comparison["runs"]) == 3
    assert len(comparison["pairwise_overlap"]) == 3

    decision = app.get_decision_summary()
    assert decision["accepted_detector"] == "MNI"
    assert decision["active_detector"] == "HIL"

    session_path = tmp_path / "cropped_session.pybrain"
    app.export_app(session_path)
    restored = HFO_App.import_app(session_path)
    restored_summaries = restored.get_run_summaries()

    assert len(restored_summaries) == 3
    assert {summary["detector_name"] for summary in restored_summaries} == {"STE", "MNI", "HIL"}
    assert restored.get_decision_summary()["accepted_detector"] == "MNI"
    assert restored.get_decision_summary()["active_detector"] == "HIL"
    assert len(restored.compare_runs()["pairwise_overlap"]) == 3


def test_spindle_backend_full_chain_smoke(tmp_path):
    if importlib.util.find_spec("torch") is None:
        pytest.skip("torch is required for real spindle classification smoke")

    recording_path = _build_spindle_recording(tmp_path)

    app = SpindleApp()
    app.set_n_jobs(2)
    app.load_edf(str(recording_path))
    app.set_filter_parameter(ParamFilterSpindle(sample_freq=app.sample_freq))
    app.filter_eeg_data()

    assert app.filtered is True
    assert app.has_filtered_data() is True
    np.testing.assert_equal(app.filter_data.shape, app.eeg_data.shape)

    app.set_filter_60()
    app.set_unfiltered_60()
    np.testing.assert_allclose(app.eeg_data, app.eeg_data_un60)

    app.set_detector(
        ParamDetector(
            ParamYASA(sample_freq=app.sample_freq, n_jobs=app.n_jobs),
            detector_type="YASA",
        )
    )
    app.detect_biomarker()

    spindle_events = app.event_features.get_num_biomarker()
    assert spindle_events >= 3
    assert np.unique(app.event_features.channel_names).tolist() == ["C3"]

    app.set_default_cpu_classifier()
    app.classify_artifacts([0.0, 0.0], 0.5)
    app.classify_spikes()

    assert app.classified is True
    assert app.event_features.has_feature() is True
    assert app.event_features.artifact_predicted is True
    assert app.event_features.spike_predicted is True
    assert len(app.event_features.artifact_predictions) == spindle_events
    assert len(app.event_features.spike_predictions) == spindle_events

    accepted_run = app.accept_active_run()
    assert accepted_run is not None
    decision = app.get_decision_summary()
    assert decision["accepted_detector"] == "YASA"
    assert decision["top_channel"]["channel_name"] == "C3"

    report_csv_path = tmp_path / "spindle_report.csv"
    app.export_report(report_csv_path)
    assert report_csv_path.exists()

    features_dir = tmp_path / "spindle_features"
    app.export_features(features_dir)
    assert (features_dir / "artifact").exists()
    assert (features_dir / "spike").exists()
    assert (features_dir / "non_spike").exists()
    assert _count_feature_images(features_dir) == spindle_events

    workbook_export_path = tmp_path / "spindle_events.xlsx"
    app.export_excel(workbook_export_path)
    workbook_events = _read_sheet_rows(workbook_export_path, "Events")
    assert len(workbook_events) == spindle_events + 1
    assert "spk-Spindle" in workbook_events[0]

    clinical_summary_path = tmp_path / "spindle_clinical_summary.xlsx"
    app.export_clinical_summary(clinical_summary_path)
    decision_rows = _read_sheet_rows(clinical_summary_path, "Decision")
    assert decision_rows[1][2] == "YASA"

    session_path = tmp_path / "spindle_session.pybrain"
    app.export_app(session_path)
    restored = SpindleApp.import_app(session_path)
    restored_summaries = restored.get_run_summaries()

    assert len(restored_summaries) == 1
    assert restored_summaries[0]["detector_name"] == "YASA"
    assert restored_summaries[0]["num_events"] == spindle_events
    assert restored.get_decision_summary()["accepted_detector"] == "YASA"
