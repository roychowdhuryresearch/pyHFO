import importlib.util
import os
from pathlib import Path

import mne
import numpy as np
import pytest
from openpyxl import load_workbook

from src.models.main_window_model import MainWindowModel
from src.param.param_detector import ParamDetector, ParamYASA
from src.param.param_filter import ParamFilterSpindle
from src.spindle_app import SpindleApp
from src.ui.annotation import Annotation
from src.ui.main_window import MainWindow
from src.ui.quick_detection import HFOQuickDetector
from src.utils.utils_detector import has_yasa


if os.environ.get("PYHFO_RUN_SMOKE") != "1":
    pytest.skip("app smoke tests are opt-in; set PYHFO_RUN_SMOKE=1 to run them", allow_module_level=True)

if importlib.util.find_spec("torch") is None:
    pytest.skip("app smoke tests require torch for real classifier execution", allow_module_level=True)


ROOT = Path(__file__).resolve().parents[1]
SAMPLE_EDF = ROOT / "SM_B_ave.edf"


@pytest.fixture(scope="session")
def hfo_cropped_fif_path(tmp_path_factory):
    raw = mne.io.read_raw_edf(SAMPLE_EDF, preload=True, verbose=False)
    raw.crop(tmin=0, tmax=40)
    raw.pick(raw.ch_names[:8])
    path = tmp_path_factory.mktemp("app_smoke") / "cropped_hfo_smoke_raw.fif"
    raw.save(path, overwrite=True, verbose=False)
    return path


@pytest.fixture(scope="session")
def hfo_midwindow_fif_path(tmp_path_factory):
    raw = mne.io.read_raw_edf(SAMPLE_EDF, preload=True, verbose=False)
    raw.crop(tmin=95, tmax=160)
    raw.pick(["POL IO3", "POL IO4", "POL IO5", "POL IO6", "POL IO7"])
    path = tmp_path_factory.mktemp("app_smoke") / "midwindow_hfo_smoke_raw.fif"
    raw.save(path, overwrite=True, verbose=False)
    return path


def _process_events(qapp, cycles=20):
    for _ in range(cycles):
        qapp.processEvents()


def _close_widget(widget, qapp):
    if widget is None:
        return
    widget.close()
    _process_events(qapp)


def _create_main_window(monkeypatch, qapp):
    monkeypatch.setattr(MainWindowModel, "init_error_terminal_display", lambda self: None)
    window = MainWindow()
    window.show()
    _process_events(qapp)
    return window


def _load_recording_into_model(model, path):
    results = model.read_edf(str(path), None)
    model.update_edf_info(results)


def _run_filter_via_model(model):
    model._filter(None)
    model.filtering_complete()


def _run_hfo_detector_via_model(model, detector_name):
    model.window.detector_mode_combo.setCurrentText(detector_name)
    if detector_name == "STE":
        model.save_ste_params()
    elif detector_name == "MNI":
        model.save_mni_params()
    else:
        raise ValueError(f"Unsupported HFO detector for smoke test: {detector_name}")
    model.backend.detect_biomarker()
    model._detect_finished()
    return model.backend.analysis_session.active_run_id


def _run_classifier_via_model(model, use_spike=True, use_ehfo=True):
    model.window.use_spike_checkbox.setChecked(use_spike)
    model.window.overview_use_spike_checkbox.setChecked(use_spike)
    if hasattr(model.window, "use_ehfo_checkbox"):
        model.window.use_ehfo_checkbox.setChecked(use_ehfo)
    if hasattr(model.window, "overview_use_ehfo_checkbox"):
        model.window.overview_use_ehfo_checkbox.setChecked(use_ehfo)
    model._classify()
    model._classify_finished()


def _find_run_row(window, detector_name):
    for row in range(window.run_table.rowCount()):
        item = window.run_table.item(row, 2)
        if item is not None and item.text() == detector_name:
            return row
    raise AssertionError(f"Could not find detector row '{detector_name}'")


def _find_annotation_window(parent_window):
    candidates = parent_window.findChildren(Annotation)
    if not candidates:
        raise AssertionError("Annotation window was not created")
    return candidates[-1]


def _read_sheet_rows(path, sheet_name):
    workbook = load_workbook(path, read_only=True, data_only=True)
    return list(workbook[sheet_name].iter_rows(values_only=True))


def _build_spindle_recording(tmp_path):
    sfreq = 2000.0
    seconds = 30
    times = np.arange(int(sfreq * seconds)) / sfreq

    signal = np.random.default_rng(0).normal(scale=5e-7, size=times.shape)
    for start in [5, 12, 20]:
        mask = (times >= start) & (times < start + 1.0)
        envelope = np.hanning(mask.sum())
        signal[mask] += 2e-5 * np.sin(2 * np.pi * 13 * times[mask]) * envelope

    reference = np.random.default_rng(1).normal(scale=4e-7, size=times.shape)
    info = mne.create_info(["C3", "C4"], sfreq, ch_types="eeg")
    raw = mne.io.RawArray(np.vstack([signal, reference]), info, verbose=False)
    raw.info["line_freq"] = 60

    path = tmp_path / "synthetic_spindle_raw_eeg.fif"
    raw.save(path, overwrite=True, verbose=False)
    return path


def _build_spindle_session(tmp_path):
    if not has_yasa():
        pytest.skip("spindle app smoke test requires the optional yasa dependency")

    recording_path = _build_spindle_recording(tmp_path)
    app = SpindleApp()
    app.set_n_jobs(2)
    app.load_edf(str(recording_path))
    app.set_filter_parameter(ParamFilterSpindle(sample_freq=app.sample_freq))
    app.filter_eeg_data()
    app.set_detector(
        ParamDetector(
            ParamYASA(sample_freq=app.sample_freq, n_jobs=app.n_jobs),
            detector_type="YASA",
        )
    )
    app.detect_biomarker()
    app.set_default_cpu_classifier()
    app.classify_artifacts([0.0, 0.0], 0.5)
    app.classify_spikes()
    app.accept_active_run()
    session_path = tmp_path / "spindle_session.pybrain"
    app.export_app(session_path)
    return session_path


def test_hfo_quick_detection_smoke_exports_excel_and_session(qapp, hfo_midwindow_fif_path):
    dialog = HFOQuickDetector()
    try:
        results = dialog.read_edf(str(hfo_midwindow_fif_path), None)
        dialog.update_edf_info(results)
        dialog.detectionTypeComboBox.setCurrentText("MNI")
        dialog.update_detector_tab("MNI")
        dialog.qd_use_classifier_checkbox.setChecked(True)
        dialog.qd_use_spikes_checkbox.setChecked(True)
        dialog.qd_use_ehfo_checkbox.setChecked(True)
        dialog.qd_ignore_sec_before_input.setText("0")
        dialog.qd_ignore_sec_after_input.setText("0")
        dialog.qd_excel_checkbox.setChecked(True)
        dialog.qd_npz_checkbox.setChecked(True)

        run_config = dialog.collect_run_configuration()
        outputs = dialog._run(run_config, None)
        dialog._run_finished(outputs)
        dialog._run_cleanup()

        assert dialog.backend.event_features.artifact_predicted is True
        assert dialog.backend.event_features.spike_predicted is True
        assert dialog.backend.event_features.ehfo_predicted is True

        assert dialog.run_button.isEnabled() is True
        assert Path(outputs["excel_output"]).exists()
        assert Path(outputs["npz_output"]).exists()

        workbook_rows = _read_sheet_rows(outputs["excel_output"], "Events")
        assert len(workbook_rows) > 1
        headers = [str(value).lower() for value in workbook_rows[0] if value is not None]
        assert "channel_names" in headers
        assert "spkhfo" in headers
        assert "ehfo" in headers

        restored = dialog.backend.import_app(outputs["npz_output"])
        assert restored.detected is True
        assert restored.classified is True
        assert len(restored.get_run_summaries()) == 1
        assert restored.event_features.get_num_biomarker() > 0
        assert restored.event_features.artifact_predicted is True
        assert restored.event_features.spike_predicted is True
        assert restored.event_features.ehfo_predicted is True
    finally:
        _close_widget(dialog, qapp)


def test_main_window_hfo_annotation_export_import_and_multi_run_smoke(monkeypatch, qapp, tmp_path, hfo_cropped_fif_path):
    window = _create_main_window(monkeypatch, qapp)
    restored_window = None
    annotation_window = None
    try:
        _load_recording_into_model(window.model, hfo_cropped_fif_path)
        _run_filter_via_model(window.model)

        ste_run_id = _run_hfo_detector_via_model(window.model, "STE")
        assert window.model.backend.event_features.get_num_biomarker() > 0

        mni_run_id = _run_hfo_detector_via_model(window.model, "MNI")
        assert mni_run_id != ste_run_id

        _run_classifier_via_model(window.model, use_spike=True, use_ehfo=True)

        assert len(window.model.backend.get_run_summaries()) == 2
        assert window.run_table.rowCount() == 2
        assert window.compare_runs_button.isEnabled() is True
        assert window.detect_all_button.isEnabled() is True

        window.model.show_run_comparison()
        assert window.comparison_table.rowCount() >= 1

        first_visibility_item = window.run_table.item(0, 0)
        first_visibility_item.setCheckState(0)
        window.model.handle_run_table_click(0, 0)
        assert len(window.model.backend.get_visible_runs()) == 1

        export_path = tmp_path / "main_window_clinical_summary.xlsx"
        monkeypatch.setattr(
            "src.models.main_window_model.QFileDialog.getSaveFileName",
            lambda *args, **kwargs: (str(export_path), "Excel files (*.xlsx)"),
        )
        window.model.save_to_excel()
        assert export_path.exists()
        assert window.model.backend.get_decision_summary()["accepted_detector"] == "MNI"

        ste_row = _find_run_row(window, "STE")
        window.model.activate_run_from_table(ste_row, 0)
        assert window.model.backend.analysis_session.active_run_id == ste_run_id
        window.model.accept_active_run()
        assert window.model.backend.get_decision_summary()["accepted_detector"] == "STE"

        window.model.open_annotation()
        _process_events(qapp)
        annotation_window = _find_annotation_window(window)
        annotation_window.select_annotation_option("Artifact")
        annotation_window.update_button_clicked()

        active_run = window.model.backend.analysis_session.get_active_run()
        assert active_run is not None
        assert active_run.event_features.annotated[0] == 1
        assert active_run.event_features.artifact_annotations[0] == 1

        snapshot_path = tmp_path / "review_snapshot.png"
        monkeypatch.setattr(
            "src.ui.annotation.QtWidgets.QFileDialog.getSaveFileName",
            lambda *args, **kwargs: (str(snapshot_path), "PNG Image (*.png)"),
        )
        annotation_window.export_snapshot()
        assert snapshot_path.exists()
        _close_widget(annotation_window, qapp)
        annotation_window = None

        session_path = tmp_path / "main_window_session.pybrain"
        window.model._save_to_npz(session_path, None)
        assert session_path.exists()

        restored_window = _create_main_window(monkeypatch, qapp)
        restored_window.model._offer_direct_annotation = lambda: None
        restored_window.model._load_from_npz(session_path, None)
        restored_window.model.load_from_npz_finished()

        restored_active = restored_window.model.backend.analysis_session.get_active_run()
        assert restored_window.model.biomarker_type == "HFO"
        assert restored_window.run_table.rowCount() == 2
        assert restored_active is not None
        assert restored_active.event_features.annotated[0] == 1
        assert restored_active.event_features.artifact_annotations[0] == 1
        assert restored_window.model.backend.get_decision_summary()["accepted_detector"] == "STE"
        assert restored_window.annotation_button.isEnabled() is True
    finally:
        _close_widget(annotation_window, qapp)
        _close_widget(restored_window, qapp)
        _close_widget(window, qapp)


def test_main_window_spindle_and_spike_branch_smoke(monkeypatch, qapp, tmp_path):
    window = _create_main_window(monkeypatch, qapp)
    annotation_window = None
    try:
        window.combo_box_biomarker.setCurrentText("Spike")
        _process_events(qapp)

        assert window.model.biomarker_type == "Spike"
        assert window.model.biomarker_supports_detection() is False
        assert window.model.biomarker_supports_classification() is False
        assert window.detector_run_button.isEnabled() is False
        assert window.classifier_run_button.isEnabled() is False
        assert window.detector_run_button.text() == "Review Only"
        assert window.classifier_run_button.text() == "Review Only"

        spindle_session_path = _build_spindle_session(tmp_path)
        window.model._offer_direct_annotation = lambda: None
        window.model._load_from_npz(spindle_session_path, None)
        window.model.load_from_npz_finished()

        assert window.model.biomarker_type == "Spindle"
        assert window.model.biomarker_supports_detection() is True
        assert window.model.biomarker_supports_classification() is True
        assert window.run_table.rowCount() == 1
        assert "Total spindles" in window.statistics_label.text()
        assert window.use_ehfo_checkbox.isVisible() is False
        assert window.overview_use_ehfo_checkbox.isVisible() is False

        export_path = tmp_path / "spindle_clinical_summary.xlsx"
        monkeypatch.setattr(
            "src.models.main_window_model.QFileDialog.getSaveFileName",
            lambda *args, **kwargs: (str(export_path), "Excel files (*.xlsx)"),
        )
        window.model.save_to_excel()
        assert export_path.exists()

        window.model.open_annotation()
        _process_events(qapp)
        annotation_window = _find_annotation_window(window)
        annotation_window.select_annotation_option("Spike")
        annotation_window.update_button_clicked()

        active_run = window.model.backend.analysis_session.get_active_run()
        assert active_run is not None
        assert active_run.event_features.annotated[0] == 1
        assert active_run.event_features.spike_annotations[0] == 1
        assert active_run.event_features.artifact_annotations[0] == 1
    finally:
        _close_widget(annotation_window, qapp)
        _close_widget(window, qapp)
